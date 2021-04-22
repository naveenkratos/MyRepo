import json
import os
import xlsxwriter
from openpyxl import load_workbook

class JsonToXlsx:

    def __init__(self, config):
        self.config = config

    def convert(self):

        print("Started JSON to Xlsx Conversion")

        #check for file exists
        xlsxFileExists = os.path.exists(self.config.XLSX_FILE_PATH)

        if(xlsxFileExists!=True):
            #create new xlsx file 
            workbook = xlsxwriter.Workbook(self.config.XLSX_FILE_PATH)
            worksheet = workbook.add_worksheet()
            workbook.close()

        wb = load_workbook(self.config.XLSX_FILE_PATH)

        sheet = wb.active

        with open(self.config.JSON_FILE_PATH) as f:
            jsonData = json.load(f)

        sheetHeadList = list(jsonData['data'][0].keys())

        #Adding Headings in every columns in sheet
        for index,head in enumerate(sheetHeadList):
            cell = sheet.cell(row = 1, column = index+1 )
            cell.value = head

        #Appending data in row
        for dataIndex,data in enumerate(jsonData['data']):
            for HeadIndex,head in enumerate(sheetHeadList):
                cell = sheet.cell(row = dataIndex+2, column = HeadIndex+1)
                cell.value = data[head]
        try:
            wb.save(self.config.XLSX_FILE_PATH)
        except:
            print("Error in saving xlsx file")
            return

        print("Finished JSON to Xlsx Conversion")